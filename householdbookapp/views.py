import json
import pandas as pd
from django.contrib.auth.models import User

from django.forms import formset_factory
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render

# Create your views here.
from django.urls import reverse
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.views.generic.edit import FormMixin
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from pandas import read_excel
from pyecharts.charts import Pie, Grid, Scatter, Line, Bar
from pyecharts import options as opts
from pyecharts.faker import Faker
from rest_framework.views import APIView

from dashboardapp.models import Dashboard
from diamond_goose.pyecharts import json_response
from exchangeapp.models import ForeignCurrency
from householdbookapp.forms import LiquidityCreationForm, DebtCreationForm, IncomeExpenseCreationForm, \
    IncomeExpenseTabularInsertForm, BuyNowPayLaterCreationForm
from householdbookapp.models import Liquidity, Debt, IncomeExpense, BuyNowPayLater
from masterinfoapp.models import CurrencyMaster
from diamond_goose.factory import format_mask_currency


class HouseholdbookHomeView(ListView):
    model = Liquidity
    template_name = 'householdbookapp/householdbook_home.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(HouseholdbookHomeView, self).get_context_data(**kwargs)

        queryset_my_dashboard = Dashboard.objects.get(owner=self.request.user)
        dashboard_pk = queryset_my_dashboard.pk
        context.update({'dashboard_pk': dashboard_pk})

        # Chart URL
        ip_address = None
        householdbook_chart_url_list = ['http://']
        try:
            from diamond_goose.settings.local import LOCAL_IP_ADDRESS
            ip_address = LOCAL_IP_ADDRESS
        except Exception as deploy_environment:
            print('Householdbook Home Chart - deploy_environment : {}'.format(deploy_environment))
            from diamond_goose.settings.deploy import DEPLOY_IP_ADDRESS
            ip_address = DEPLOY_IP_ADDRESS

        if ip_address:
            householdbook_chart_url_list.append(ip_address)
            householdbook_chart_url_list.append('/householdbook/householdbook_chart/')
            context.update({'householdbook_chart_url_list': ''.join(householdbook_chart_url_list)})

        return context


class HouseholdbookChartView(APIView):
    def get(self, request, *args, **kwargs):

        chart_element_position = {
            'liquidity_pie_chart_position': ["19%", "60%"],
            'liquidity_title_position_left': "0%",
            'liquidity_legend_position_left': "0%",

            'debt_pie_chart_position': ["50%", "60%"],
            'debt_title_position_left': "32%",
            'debt_legend_position_left': "32%",

            'title_font_size': 20,
        }


        liquidity_pie_chart_data = liquidity_pie_chart_data_generator(request)
        debt_pie_chart_data = debt_pie_chart_data_generator(request)
        for color in debt_pie_chart_data['color_list']:
            liquidity_pie_chart_data['color_list'].append(color)

        liquidity_pie_chart = Pie()
        liquidity_pie_chart.add(
            series_name="Liquidity Composition",
            data_pair=liquidity_pie_chart_data['data_pair'],
            radius=["40%", "70%"],
            # rosetype="radius",
            center=chart_element_position['liquidity_pie_chart_position'],
            label_opts=opts.LabelOpts(is_show=False, position="center"),
        )
        liquidity_pie_chart.set_colors(
            liquidity_pie_chart_data['color_list']
        )

        liquidity_pie_chart.set_global_opts(
            title_opts=opts.TitleOpts(
                title=liquidity_pie_chart_data['total_amount_text'],
                pos_left=chart_element_position['liquidity_title_position_left'],
                pos_top="5",
                title_textstyle_opts=opts.TextStyleOpts(color="#FFFFFF",
                                                        font_size=chart_element_position['title_font_size']),
            ),
            legend_opts=opts.LegendOpts(pos_left=chart_element_position['liquidity_legend_position_left'],
                                        pos_top="center",
                                        orient="vertical",
                                        textstyle_opts=opts.TextStyleOpts(color="#FFFFFF")),
        )
        liquidity_pie_chart.set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(color="#FFFFFF"),
        )


        print(debt_pie_chart_data['color_list'])
        debt_pie_chart = Pie()
        debt_pie_chart.add(
            series_name="Liquidity Composition",
            data_pair=debt_pie_chart_data['data_pair'],
            radius=["40%", "70%"],
            # rosetype="radius",
            center=chart_element_position['debt_pie_chart_position'],
            label_opts=opts.LabelOpts(is_show=False, position="center"),
        )

        debt_pie_chart.set_colors(
            debt_pie_chart_data['color_list']
        )

        debt_pie_chart.set_global_opts(
            title_opts=opts.TitleOpts(
                title=debt_pie_chart_data['total_amount_text'],
                pos_left=chart_element_position['debt_title_position_left'],
                pos_top="5",
                title_textstyle_opts=opts.TextStyleOpts(color="#FFFFFF",
                                                        font_size=chart_element_position['title_font_size']),
            ),
            legend_opts=opts.LegendOpts(pos_left=chart_element_position['debt_legend_position_left'],
                                        pos_top="center",
                                        orient="vertical",
                                        textstyle_opts=opts.TextStyleOpts(color="#FFFFFF")),
        )

        debt_pie_chart.set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(color="#FFFFFF"),
        )

        grid = (
            Grid()
            .add(
                liquidity_pie_chart,
                grid_opts=opts.GridOpts(pos_left="55%")
            )
            .add(
                debt_pie_chart,
                grid_opts=opts.GridOpts(pos_right="55%")
            )
            .dump_options_with_quotes()
        )

        return json_response(json.loads(grid))


class LiquidityListView(ListView):
    model = Liquidity
    template_name = 'householdbookapp/liquidity_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(LiquidityListView, self).get_context_data(**kwargs)

        queryset_my_dashboard = Dashboard.objects.get(owner=self.request.user)
        dashboard_pk = queryset_my_dashboard.pk
        context.update({'dashboard_pk': dashboard_pk})

        queryset_my_liquidities = Liquidity.objects.filter(owner=self.request.user,
                                                           dashboard=dashboard_pk)
        for liquidity in queryset_my_liquidities:
            liquidity.update_statistics()
        total_liquidity_amount = 0

        for liquidity in queryset_my_liquidities:
            # liquidity.update_statistics()
            total_liquidity_amount += liquidity.amount_exchanged
        context.update({'queryset_my_liquidities': queryset_my_liquidities})
        context.update({'total_liquidity_amount': total_liquidity_amount})
        context.update({'main_currency_code': queryset_my_dashboard.main_currency.currency_code})

        liquidity_pie_chart_url_list = ['http://']
        ip_address = None
        try:
            from diamond_goose.settings.local import LOCAL_IP_ADDRESS
            ip_address = LOCAL_IP_ADDRESS
        except Exception as deploy_environment:
            print('Liquidity Pie Chart - deploy_environment : {}'.format(deploy_environment))
            from diamond_goose.settings.deploy import DEPLOY_IP_ADDRESS
            ip_address = DEPLOY_IP_ADDRESS

        if ip_address:
            liquidity_pie_chart_url_list.append(ip_address)
            liquidity_pie_chart_url_list.append('/householdbook/liquidity_pie_chart/')
            context.update({'liquidity_pie_chart_url_list': ''.join(liquidity_pie_chart_url_list)})

        return context


class LiquidityCreateView(CreateView):
    model = Liquidity
    form_class = LiquidityCreationForm
    template_name = 'householdbookapp/liquidity_create.html'

    def form_valid(self, form):
        temp_liquidity = form.save(commit=False)
        temp_liquidity.owner = self.request.user
        temp_liquidity.dashboard = Dashboard.objects.get(owner=self.request.user)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('householdbookapp:liquidity_list')


class LiquidityUpdateView(UpdateView):
    model = Liquidity
    form_class = LiquidityCreationForm
    template_name = 'householdbookapp/liquidity_update.html'
    context_object_name = 'target_liquidity'

    def get_success_url(self):
        return reverse('householdbookapp:liquidity_list')


class LiquidityDeleteView(DeleteView):
    model = Liquidity
    template_name = 'householdbookapp/liquidity_delete.html'
    context_object_name = 'target_liquidity'

    def get_success_url(self):
        return reverse('householdbookapp:liquidity_list')


def color_list_generator(starting_hex, ending_hex, color_number):
    if color_number == 0:
        result_list = []
    elif color_number == 1:
        result_list = [starting_hex]
    elif color_number == 2:
        result_list = [starting_hex, ending_hex]
    else:
        result_list = ['#' for i in range(color_number)]
        for i in range(3):
            starting = int(starting_hex[2*i+1:2*i+3], 16)
            ending = int(ending_hex[2*i+1:2*i+3], 16)
            delta = (ending-starting)/(color_number-1)
            # print('[in hex] starting : {} / ending : {}'.format(starting_hex[2*i+1:2*i+3], ending_hex[2*i+1:2*i+1+3]))
            # print('[in decimal] starting : {} / ending : {} / delta : {}'.format(starting, ending, delta))
            for j in range(color_number):
                hex_str = str(hex(round(starting+delta*j)))[2:].upper()
                if len(hex_str) < 2:
                    new_hex_str = '0'
                    new_hex_str += hex_str
                    hex_str = new_hex_str
                # print(hex_str)
                result_list[j] += hex_str
    # print(result_list)
    return result_list


def currency_format(amount, currency_master):
    result_text_list = [currency_master.currency_sign, ' ']
    below_period = None
    if currency_master.currency_code != 'KRW':
        below_period = str(round(amount, 2)).split('.')[-1]

    integer_list = []
    while amount >= 1000:
        temp_num_str = str(round(amount%1000))
        while len(temp_num_str) < 3:
            temp_num_str = '0' + temp_num_str
        integer_list.append(temp_num_str)
        amount /= 1000
    integer_list.append(str(round(amount)))

    for i in range(len(integer_list)):
        result_text_list.append(integer_list[(i+1)*(-1)])
        result_text_list.append(',')
    result_text_list.pop(-1)
    if below_period:
        result_text_list.append('.')
        result_text_list.append(below_period)
    return ''.join(result_text_list)


def liquidity_pie_chart_data_generator(request):
    queryset_liquidity = Liquidity.objects.filter(owner=request.user)
    x_data = []
    y_data = []
    total_amount = 0
    for liquidity in queryset_liquidity:
        x_data.append(liquidity.liquidity_name)
        y_data.append(liquidity.amount_exchanged)
        total_amount += liquidity.amount_exchanged
    data_pair = [list(z) for z in zip(x_data, y_data)]
    data_pair.sort(key=lambda x: x[1])

    # color : #002921 ~ #00FACC
    color_list = color_list_generator('#002921', '#00FACC', queryset_liquidity.count())

    # total amount
    queryset_dashboard = Dashboard.objects.get(owner=request.user)
    total_amount_text = ''.join(['Total Amount : ',
                                currency_format(total_amount, queryset_dashboard.main_currency)])
    return {
        'data_pair': data_pair,
        'color_list': color_list,
        'total_amount_text': total_amount_text,
    }


def liquidity_pie_chart(request, dump_option=False) -> Pie:
    liquidity_pie_chart_data = liquidity_pie_chart_data_generator(request)
    pie_chart = Pie()
    pie_chart.add(
            series_name="Liquidity Composition",
            data_pair=liquidity_pie_chart_data['data_pair'],
            radius=["40%", "70%"],
            # rosetype="radius",
            center=["60%", "60%"],
            label_opts=opts.LabelOpts(is_show=False, position="center"),
        )
    pie_chart.set_colors(
            liquidity_pie_chart_data['color_list']
        )

    pie_chart.set_global_opts(
            title_opts=opts.TitleOpts(
                title=liquidity_pie_chart_data['total_amount_text'],
                pos_left="left",
                pos_top="5",
                title_textstyle_opts=opts.TextStyleOpts(color="#FFFFFF", font_size=20),
            ),
            legend_opts=opts.LegendOpts(pos_left="left",
                                        pos_top="center",
                                        orient="vertical",
                                        textstyle_opts=opts.TextStyleOpts(color="#FFFFFF")),
        )
    pie_chart.set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(color="#FFFFFF"),
        )
    dump_pie_chart = pie_chart.dump_options_with_quotes()

    if dump_option:
        return dump_pie_chart
    else:
        return pie_chart


class LiquidityPieChartView(APIView):
    def get(self, request, *args, **kwargs):
        return json_response(json.loads(liquidity_pie_chart(request, dump_option=True)))


class DebtListView(ListView):
    model = Debt
    template_name = 'householdbookapp/debt_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(DebtListView, self).get_context_data(**kwargs)

        queryset_my_dashboard = Dashboard.objects.get(owner=self.request.user)
        dashboard_pk = queryset_my_dashboard.pk
        context.update({'dashboard_pk': dashboard_pk})

        queryset_my_debts = Debt.objects.filter(owner=self.request.user,
                                                dashboard=dashboard_pk)
        total_debt_amount = 0

        for debt in queryset_my_debts:
            debt.update_statistics()
            total_debt_amount += debt.amount_exchanged
            if debt.long_term_debt_flag:
                debt.debt_term = 'Long'
            else:
                debt.debt_term = 'Short'
        context.update({'queryset_my_debts': queryset_my_debts})
        context.update({'total_debt_amount': total_debt_amount})
        context.update({'main_currency_code': queryset_my_dashboard.main_currency.currency_code})

        debt_pie_chart_url_list = ['http://']
        ip_address = None
        try:
            from diamond_goose.settings.local import LOCAL_IP_ADDRESS
            ip_address = LOCAL_IP_ADDRESS
        except Exception as deploy_environment:
            print('Debt Pie Chart - deploy_environment : {}'.format(deploy_environment))
            from diamond_goose.settings.deploy import DEPLOY_IP_ADDRESS
            ip_address = DEPLOY_IP_ADDRESS

        if ip_address:
            debt_pie_chart_url_list.append(ip_address)
            debt_pie_chart_url_list.append('/householdbook/debt_pie_chart/')
            context.update({'debt_pie_chart_url_list': ''.join(debt_pie_chart_url_list)})

        queryset_bnpls = BuyNowPayLater.objects.filter(owner=self.request.user,
                                                       dashboard=dashboard_pk,
                                                       end_flag=False).order_by('end_flag',
                                                                                'purchase_period')
        bnpl_summary = {
            'total_amount': 0,
            'discount_amount': 0,
            'real_remaining_amount': 0,
            'real_monthly_payment_amount': 0,
        }
        for bnpl in queryset_bnpls:
            bnpl.update_statistics()
            if not bnpl.end_flag:
                bnpl_summary['total_amount'] += bnpl.total_amount
                bnpl_summary['discount_amount'] += bnpl.discount_amount
                bnpl_summary['real_remaining_amount'] += bnpl.real_remaining_amount
                bnpl_summary['real_monthly_payment_amount'] += bnpl.real_monthly_payment_amount

            item_name_length = len(bnpl.item_name.encode('utf-8'))
            if item_name_length > 16:
                if item_name_length != len(bnpl.item_name):
                    bnpl.item_name = bnpl.item_name[0:12] + '...'
                else:
                    bnpl.item_name = bnpl.item_name[0:17] + '...'
            bnpl.payment_count = str(round(bnpl.current_payment_count))+' / '+str(round(bnpl.paying_months))
            bnpl.total_amount = format_mask_currency(bnpl.total_amount, bnpl.currency)
            bnpl.discount_amount = format_mask_currency(bnpl.discount_amount, bnpl.currency)
            bnpl.real_remaining_amount = format_mask_currency(bnpl.real_remaining_amount, bnpl.currency)
            bnpl.real_monthly_payment_amount = format_mask_currency(bnpl.real_monthly_payment_amount, bnpl.currency)
            bnpl.end_flag = 'Y' if bnpl.end_flag else 'N'

        context.update({'queryset_bnpls': queryset_bnpls})

        bnpl_summary['total_amount'] = format_mask_currency(bnpl_summary['total_amount'], queryset_my_dashboard.main_currency)
        bnpl_summary['discount_amount'] = format_mask_currency(bnpl_summary['discount_amount'], queryset_my_dashboard.main_currency)
        bnpl_summary['real_remaining_amount'] = format_mask_currency(bnpl_summary['real_remaining_amount'], queryset_my_dashboard.main_currency)
        bnpl_summary['real_monthly_payment_amount'] = format_mask_currency(bnpl_summary['real_monthly_payment_amount'], queryset_my_dashboard.main_currency)
        context.update({'bnpl_summary': bnpl_summary})

        return context


class DebtCreateView(CreateView):
    model = Debt
    form_class = DebtCreationForm
    template_name = 'householdbookapp/debt_create.html'

    def form_valid(self, form):
        temp_debt = form.save(commit=False)
        temp_debt.owner = self.request.user
        temp_debt.dashboard = Dashboard.objects.get(owner=self.request.user)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('householdbookapp:debt_list')


class DebtUpdateView(UpdateView):
    model = Debt
    form_class = DebtCreationForm
    template_name = 'householdbookapp/debt_update.html'
    context_object_name = 'target_debt'

    def get_success_url(self):
        return reverse('householdbookapp:debt_list')


class DebtDeleteView(DeleteView):
    model = Debt
    template_name = 'householdbookapp/debt_delete.html'
    context_object_name = 'target_debt'

    def get_success_url(self):
        return reverse('householdbookapp:debt_list')


def debt_pie_chart_data_generator(request):
    queryset_debt = Debt.objects.filter(owner=request.user)
    x_data = []
    y_data = []
    total_amount = 0
    for debt in queryset_debt:
        x_data.append(debt.debt_name)
        y_data.append(debt.amount_exchanged)
        total_amount += debt.amount_exchanged
    data_pair = [list(z) for z in zip(x_data, y_data)]
    data_pair.sort(key=lambda x: x[1])

    # color : #6B451A ~ #F7A03B
    color_list = color_list_generator('#6B451A', '#F7A03B', queryset_debt.count())

    # total amount
    queryset_dashboard = Dashboard.objects.get(owner=request.user)
    total_amount_text = ''.join(['Total Amount : ',
                                currency_format(total_amount, queryset_dashboard.main_currency)])

    return {
        'data_pair': data_pair,
        'color_list': color_list,
        'total_amount_text': total_amount_text,
    }


def debt_pie_chart(request, dump_option=False) -> Pie:
    debt_pie_chart_data = debt_pie_chart_data_generator(request)
    pie_chart = Pie()
    pie_chart.add(
            series_name="Liquidity Composition",
            data_pair=debt_pie_chart_data['data_pair'],
            radius=["40%", "70%"],
            # rosetype="radius",
            center=["60%", "60%"],
            label_opts=opts.LabelOpts(is_show=False, position="center"),
        )

    pie_chart.set_colors(
            debt_pie_chart_data['color_list']
        )

    pie_chart.set_global_opts(
            title_opts=opts.TitleOpts(
                title=debt_pie_chart_data['total_amount_text'],
                pos_left="left",
                pos_top="5",
                title_textstyle_opts=opts.TextStyleOpts(color="#FFFFFF", font_size=20),
            ),
            legend_opts=opts.LegendOpts(pos_left="left",
                                        pos_top="center",
                                        orient="vertical",
                                        textstyle_opts=opts.TextStyleOpts(color="#FFFFFF")),
        )

    pie_chart.set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(color="#FFFFFF"),
        )

    dump_pie_chart = pie_chart.dump_options_with_quotes()

    if dump_option:
        return dump_pie_chart
    else:
        return pie_chart


class DebtPieChartView(APIView):
    def get(self, request, *args, **kwargs):
        return json_response(json.loads(debt_pie_chart(request, dump_option=True)))


class IncomeExpenseListView(ListView, FormMixin):
    model = IncomeExpense
    form_class = IncomeExpenseCreationForm
    template_name = 'householdbookapp/income_expense_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(IncomeExpenseListView, self).get_context_data(**kwargs)

        queryset_my_dashboard = Dashboard.objects.get(owner=self.request.user)
        dashboard_pk = queryset_my_dashboard.pk
        context.update({'dashboard_pk': dashboard_pk})

        queryset_my_income_expenses = IncomeExpense.objects.filter(owner=self.request.user,
                                                                   dashboard=dashboard_pk).order_by('-period_name')
        for income_expense in queryset_my_income_expenses:
            if (income_expense.total_income_amount == 0 or income_expense.total_expense_amount == 0 or
                income_expense.total_savings_amount == 0 or income_expense.monthly_savings_rate == 0):
                income_expense.update_statistics()
                income_expense.refresh_from_db()
        context.update({'queryset_my_income_expenses': queryset_my_income_expenses})

        # Column width data
        amount_column_count = 20
        amount_column_width = 140
        table_width = sum([
            40, # Edit
            20, # Delete
            70, # Period
            110, # Savings Rate
        ])
        table_width += amount_column_count * amount_column_width
        context.update({
            'table_width_px': str(table_width)+'px',
            'amount_column_width_px': str(amount_column_width)+'px',
        })

        # Chart URL
        grid_chart_url_list = ['http://']
        ip_address = None
        try:
            from diamond_goose.settings.local import LOCAL_IP_ADDRESS
            ip_address = LOCAL_IP_ADDRESS
        except Exception as deploy_environment:
            print('Income/Expense Grid Chart - deploy_environment : {}'.format(deploy_environment))
            from diamond_goose.settings.deploy import DEPLOY_IP_ADDRESS
            ip_address = DEPLOY_IP_ADDRESS

        if ip_address:
            grid_chart_url_list.append(ip_address)
            grid_chart_url_list.append('/householdbook/income_expense_grid_chart/')
            context.update({'grid_chart_url_list': ''.join(grid_chart_url_list)})
            print(''.join(grid_chart_url_list))

        tabular_form_set_factory = formset_factory(IncomeExpenseTabularInsertForm)
        tabular_form_set = tabular_form_set_factory()
        context.update({'tabular_form_set': tabular_form_set})


        return context


class IncomeExpenseCreateView(CreateView):
    model = IncomeExpense
    form_class = IncomeExpenseCreationForm
    template_name = 'householdbookapp/income_expense_create.html'

    def form_valid(self, form):
        temp_income_expense = form.save(commit=False)
        temp_income_expense.owner = self.request.user
        temp_income_expense.dashboard = Dashboard.objects.get(owner=self.request.user)
        temp_income_expense.save()

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('householdbookapp:income_expense_list')


class IncomeExpenseTabularInsert(CreateView):
    model = IncomeExpense
    form_class = IncomeExpenseTabularInsertForm
    template_name = 'householdbookapp/income_expense_tabular_input.html'

    def form_valid(self, form):
        temp_income_expense = form.save(commit=False)
        temp_income_expense.owner = self.request.user
        temp_income_expense.dashboard = Dashboard.objects.get(owner=self.request.user)
        print(self.request)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('householdbookapp:income_expense_list')


class IncomeExpenseUpdateView(UpdateView):
    model = IncomeExpense
    form_class = IncomeExpenseCreationForm
    template_name = 'householdbookapp/income_expense_update.html'
    context_object_name = 'target_income_expense'

    def get_success_url(self):
        return reverse('householdbookapp:income_expense_list')


class IncomeExpenseDeleteView(DeleteView):
    model = IncomeExpense
    template_name = 'householdbookapp/income_expense_delete.html'
    context_object_name = 'target_income_expense'

    def get_success_url(self):
        return reverse('householdbookapp:income_expense_list')


def income_expense_delete_all(request):
    dashboard_pk = request.GET['dashboard_pk']
    queryset_asset_income_expenses = IncomeExpense.objects.filter(owner=request.user,
                                                                  dashboard=dashboard_pk)
    delete_count = 0
    deleted_periods_list = []
    for income_expense in queryset_asset_income_expenses:
        delete_count += 1
        deleted_periods_list.append(income_expense.period_name)
        income_expense.delete()
    print('Delete Income/Expense : {}'.format(','.join(deleted_periods_list)))

    return HttpResponseRedirect(reverse('householdbookapp:income_expense_list'))


def income_expense_excel_download(request):
    if request.method == 'POST':
        dashboard_pk = request.POST['dashboard_pk']
    queryset_income_expense = IncomeExpense.objects.filter(dashboard=dashboard_pk).order_by('period_name')

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "transactions"
        output_row_count = 1

        # Header Insert
        worksheet.cell(row=output_row_count, column=1, value="period_name")
        worksheet.cell(row=output_row_count, column=2, value="income_labor")
        worksheet.cell(row=output_row_count, column=3, value="income_capital")
        worksheet.cell(row=output_row_count, column=4, value="income_etc")
        worksheet.cell(row=output_row_count, column=5, value="expense_housing_communication")
        worksheet.cell(row=output_row_count, column=6, value="expense_living")
        worksheet.cell(row=output_row_count, column=7, value="expense_grocery")
        worksheet.cell(row=output_row_count, column=8, value="expense_shopping")
        worksheet.cell(row=output_row_count, column=9, value="expense_leisure")
        worksheet.cell(row=output_row_count, column=10, value="expense_health")
        worksheet.cell(row=output_row_count, column=11, value="expense_largess")
        worksheet.cell(row=output_row_count, column=12, value="expense_transportation")
        worksheet.cell(row=output_row_count, column=13, value="expense_alcohol")
        worksheet.cell(row=output_row_count, column=14, value="expense_education")
        worksheet.cell(row=output_row_count, column=15, value="expense_car")
        worksheet.cell(row=output_row_count, column=16, value="expense_financial")
        worksheet.cell(row=output_row_count, column=17, value="expense_love_affair")
        worksheet.cell(row=output_row_count, column=18, value="expense_etc")

        # Line Insert
        for income_expense in queryset_income_expense:
            output_row_count += 1
            worksheet.cell(row=output_row_count, column=1, value=income_expense.period_name)
            worksheet.cell(row=output_row_count, column=2, value=income_expense.income_labor)
            worksheet.cell(row=output_row_count, column=3, value=income_expense.income_capital)
            worksheet.cell(row=output_row_count, column=4, value=income_expense.income_etc)
            worksheet.cell(row=output_row_count, column=5, value=income_expense.expense_housing_communication)
            worksheet.cell(row=output_row_count, column=6, value=income_expense.expense_living)
            worksheet.cell(row=output_row_count, column=7, value=income_expense.expense_grocery)
            worksheet.cell(row=output_row_count, column=8, value=income_expense.expense_shopping)
            worksheet.cell(row=output_row_count, column=9, value=income_expense.expense_leisure)
            worksheet.cell(row=output_row_count, column=10, value=income_expense.expense_health)
            worksheet.cell(row=output_row_count, column=11, value=income_expense.expense_largess)
            worksheet.cell(row=output_row_count, column=12, value=income_expense.expense_transportation)
            worksheet.cell(row=output_row_count, column=13, value=income_expense.expense_alcohol)
            worksheet.cell(row=output_row_count, column=14, value=income_expense.expense_education)
            worksheet.cell(row=output_row_count, column=15, value=income_expense.expense_car)
            worksheet.cell(row=output_row_count, column=16, value=income_expense.expense_financial)
            worksheet.cell(row=output_row_count, column=17, value=income_expense.expense_love_affair)
            worksheet.cell(row=output_row_count, column=18, value=income_expense.expense_etc)

        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=income_expense.xlsx'
        return response

    except Exception as identifier:
        print('income_expense_excel_download: excel_export', identifier)


class IncomeExpenseExcelUploadButton(ListView):
    model = IncomeExpense
    template_name = 'householdbookapp/income_expense_excel_upload.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(IncomeExpenseExcelUploadButton, self).get_context_data(**kwargs)
        context.update({
            'user_id': self.request.user.id,
            'dashboard_pk': Dashboard.objects.get(owner=self.request.user).pk,
        })

        return context


def income_expense_excel_upload(request):
    try:
        if request.method == 'POST' and request.FILES['transaction_file']:
            db_frame = pd.read_excel(request.FILES['transaction_file'], sheet_name=0).fillna('')
            target_dashboard = Dashboard.objects.get(pk=request.POST['dashboard_pk'])
    except Exception as identifier:
        print('income_expense_excel_upload: excel_import', identifier)

    try:
        transaction_object_list = []
        for row in db_frame.itertuples():
            upload_format = {}
            upload_format['period_name'] = row[1]
            upload_format['income_labor'] = row[2]
            upload_format['income_capital'] = row[3]
            upload_format['income_etc'] = row[4]
            upload_format['expense_housing_communication'] = row[5]
            upload_format['expense_living'] = row[6]
            upload_format['expense_grocery'] = row[7]
            upload_format['expense_shopping'] = row[8]
            upload_format['expense_leisure'] = row[9]
            upload_format['expense_health'] = row[10]
            upload_format['expense_largess'] = row[11]
            upload_format['expense_transportation'] = row[12]
            upload_format['expense_alcohol'] = row[13]
            upload_format['expense_education'] = row[14]
            upload_format['expense_car'] = row[15]
            upload_format['expense_financial'] = row[16]
            upload_format['expense_love_affair'] = row[17]
            upload_format['expense_etc'] = row[18]

            upload_format['total_income_amount'] = sum([
                upload_format['income_labor'],
                upload_format['income_capital'],
                upload_format['income_etc'],
            ])

            upload_format['total_expense_amount'] = sum([
                upload_format['expense_housing_communication'],
                upload_format['expense_living'],
                upload_format['expense_grocery'],
                upload_format['expense_shopping'],
                upload_format['expense_leisure'],
                upload_format['expense_health'],
                upload_format['expense_largess'],
                upload_format['expense_transportation'],
                upload_format['expense_alcohol'],
                upload_format['expense_education'],
                upload_format['expense_car'],
                upload_format['expense_financial'],
                upload_format['expense_love_affair'],
                upload_format['expense_etc'],
            ])

            upload_format['total_savings_amount'] = upload_format['total_income_amount'] - upload_format['total_expense_amount']

            if upload_format['total_income_amount'] > 0:
                upload_format['monthly_savings_rate'] = upload_format['total_savings_amount'] / upload_format['total_income_amount']
            else:
                upload_format['monthly_savings_rate'] = 0

            # print(upload_format)

            obj = IncomeExpense.objects.create(
                owner=request.user,
                dashboard=target_dashboard,
                period_name=upload_format['period_name'],
                income_labor=upload_format['income_labor'],
                income_capital=upload_format['income_capital'],
                income_etc=upload_format['income_etc'],
                expense_housing_communication=upload_format['expense_housing_communication'],
                expense_living=upload_format['expense_living'],
                expense_grocery=upload_format['expense_grocery'],
                expense_shopping=upload_format['expense_shopping'],
                expense_leisure=upload_format['expense_leisure'],
                expense_health=upload_format['expense_health'],
                expense_largess=upload_format['expense_largess'],
                expense_transportation=upload_format['expense_transportation'],
                expense_alcohol=upload_format['expense_alcohol'],
                expense_education=upload_format['expense_education'],
                expense_car=upload_format['expense_car'],
                expense_financial=upload_format['expense_financial'],
                expense_love_affair=upload_format['expense_love_affair'],
                expense_etc=upload_format['expense_etc'],
                total_income_amount=upload_format['total_income_amount'],
                total_expense_amount=upload_format['total_expense_amount'],
                total_savings_amount=upload_format['total_savings_amount'],
                monthly_savings_rate=upload_format['monthly_savings_rate'],
            )

            transaction_object_list.append(obj)

        # Transaction COMMIT
        for transaction in transaction_object_list:
            try:
                transaction.save()
                print('Transaction Added :', transaction)
            except Exception as identifier:
                print('Exception in saving transaction.', identifier, transaction)
                continue

    except Exception as identifier:
        print('income_expense_excel_upload: Transaction INSERT :', identifier)

    return HttpResponseRedirect(reverse('householdbookapp:income_expense_list'))


def income_expense_chart_grid_data_generator(request):
    queryset_income_expenses = IncomeExpense.objects.filter(owner=request.user).order_by('period_name')
    x_data = []
    bar_y_data = [
        {'name': 'Income',  'data_set': []},
        {'name': 'Expense', 'data_set': []},
    ]
    line_y_data = [
        {'name': 'Housing/Comm.', 'data_set': []},
        {'name': 'Living', 'data_set': []},
        {'name': 'Grocery', 'data_set': []},
        {'name': 'Shopping', 'data_set': []},
        {'name': 'Leisure', 'data_set': []},
        {'name': 'Health', 'data_set': []},
        {'name': 'Largess', 'data_set': []},
        {'name': 'Transport.', 'data_set': []},
        {'name': 'Alcohol', 'data_set': []},
        {'name': 'Edu.', 'data_set': []},
        {'name': 'Car', 'data_set': []},
        {'name': 'Fin.', 'data_set': []},
        {'name': '♥', 'data_set': []},
        {'name': 'etc', 'data_set': []},
    ]

    max_income_expense = 0
    max_among_expenses = 0

    for income_expense in queryset_income_expenses:
        x_data.append(income_expense.period_name)
        max_income_expense = max(max_income_expense, max(income_expense.total_expense_amount,
                                                         income_expense.total_income_amount))
        max_among_expenses = max(
            max_among_expenses,
            max([
                income_expense.expense_housing_communication,
                income_expense.expense_living,
                income_expense.expense_grocery,
                income_expense.expense_shopping,
                income_expense.expense_leisure,
                income_expense.expense_health,
                income_expense.expense_largess,
                income_expense.expense_transportation,
                income_expense.expense_alcohol,
                income_expense.expense_education,
                income_expense.expense_car,
                income_expense.expense_financial,
                income_expense.expense_love_affair,
                income_expense.expense_etc,
            ]))

        for element in bar_y_data:
            if element['name'] =='Income' :     element['data_set'].append(income_expense.total_income_amount)
            if element['name'] =='Expense':     element['data_set'].append(income_expense.total_expense_amount)

        for expense in line_y_data:
            if expense['name'] =='Housing/Comm.':     expense['data_set'].append(income_expense.expense_housing_communication)
            if expense['name'] =='Living'       :     expense['data_set'].append(income_expense.expense_living)
            if expense['name'] =='Grocery'      :     expense['data_set'].append(income_expense.expense_grocery)
            if expense['name'] =='Shopping'     :     expense['data_set'].append(income_expense.expense_shopping)
            if expense['name'] =='Leisure'      :     expense['data_set'].append(income_expense.expense_leisure)
            if expense['name'] =='Health'       :     expense['data_set'].append(income_expense.expense_health)
            if expense['name'] =='Largess'      :     expense['data_set'].append(income_expense.expense_largess)
            if expense['name'] =='Transport.'   :     expense['data_set'].append(income_expense.expense_transportation)
            if expense['name'] =='Alcohol'      :     expense['data_set'].append(income_expense.expense_alcohol)
            if expense['name'] =='Edu.'         :     expense['data_set'].append(income_expense.expense_education)
            if expense['name'] =='Car'          :     expense['data_set'].append(income_expense.expense_car)
            if expense['name'] =='Fin.'         :     expense['data_set'].append(income_expense.expense_financial)
            if expense['name'] =='♥'             :    expense['data_set'].append(income_expense.expense_love_affair)
            if expense['name'] =='etc'          :     expense['data_set'].append(income_expense.expense_etc)

    return {
        'x_data': x_data,
        'bar_y_data': bar_y_data,
        'line_y_data': line_y_data,
        'max_income_expense': round(max_income_expense * 1.1),
        'max_among_expenses': round(max_among_expenses * 0.8),
    }


class IncomeExpenseGridChartView(APIView):
    def get(self, request, *args, **kwargs):

        chart_base_data = income_expense_chart_grid_data_generator(self.request)
        bar_graph = Bar()
        line_graph = Line()

        bar_graph.add_xaxis(xaxis_data=chart_base_data['x_data'])
        for element in chart_base_data['bar_y_data']:
            bar_graph.add_yaxis(
                series_name=element['name'],
                y_axis=element['data_set'],
                yaxis_index=1,
                bar_width=10,
                label_opts=opts.LabelOpts(is_show=False),
                itemstyle_opts=opts.ItemStyleOpts(opacity=50)
            )

        bar_graph.extend_axis(
            yaxis=opts.AxisOpts(
                name="Income/Expense",
                type_="value",
                min_=0,
                max_=chart_base_data['max_income_expense'],
                position="left",
                axisline_opts=opts.AxisLineOpts(
                    linestyle_opts=opts.LineStyleOpts(color="#FFFFFF")),
                axislabel_opts=opts.LabelOpts(formatter="{value}"),
                axispointer_opts=opts.AxisPointerOpts(is_show=True, label=opts.LabelOpts(color="#081321",
                                                                                         font_weight='bold'))
            )
        )

        bar_graph.extend_axis(
            yaxis=opts.AxisOpts(
                name="Expenses",
                type_="value",
                min_=0,
                max_=chart_base_data['max_among_expenses'],
                position="right",
                axisline_opts=opts.AxisLineOpts(
                    linestyle_opts=opts.LineStyleOpts(color="#FFFFFF")
                ),
                axislabel_opts=opts.LabelOpts(formatter="{value}"),
                axispointer_opts=opts.AxisPointerOpts(is_show=True, label=opts.LabelOpts(color="#081321",
                                                                                         font_weight='bold'))
            )
        )

        bar_graph.set_global_opts(
            title_opts=opts.TitleOpts(title="Income Expense Summary",
                                      title_textstyle_opts=opts.TextStyleOpts(color="#FFFFFF")),
            tooltip_opts=opts.TooltipOpts(trigger="axis",
                                          axis_pointer_type="cross"),
            xaxis_opts=opts.AxisOpts(axisline_opts=opts.AxisLineOpts(linestyle_opts=opts.LineStyleOpts(color="#FFFFFF"))),
            yaxis_opts=opts.AxisOpts(axispointer_opts=opts.AxisPointerOpts(label=opts.LabelOpts(color="#081321",
                                                                                                font_weight='bold'))),
            legend_opts=opts.LegendOpts(textstyle_opts=opts.TextStyleOpts(color="#FFFFFF"))
        )

        line_graph.add_xaxis(chart_base_data['x_data'])
        for expense in chart_base_data['line_y_data']:
            line_graph.add_yaxis(
                series_name=expense['name'],
                y_axis=expense['data_set'],
                yaxis_index=2,
                is_symbol_show=False,
                linestyle_opts=opts.LineStyleOpts(type_='solid', width=2.5),
            )
        line_graph.set_global_opts(
            legend_opts=opts.LegendOpts(textstyle_opts=opts.TextStyleOpts(color="#FFFFFF"))
        )

        bar_graph.overlap(line_graph)
        bar_graph_dump = bar_graph.dump_options_with_quotes()

        return json_response(json.loads(bar_graph_dump))


class BuyNowPayLaterCreateView(CreateView):
    model = BuyNowPayLater
    form_class = BuyNowPayLaterCreationForm
    template_name = 'householdbookapp/bnpl_create.html'

    def form_valid(self, form):
        temp_bnpl = form.save(commit=False)
        temp_bnpl.owner = self.request.user
        temp_bnpl.dashboard = Dashboard.objects.get(owner=self.request.user)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('householdbookapp:debt_list')


class BuyNowPayLaterDeleteView(DeleteView):
    model = BuyNowPayLater
    template_name = 'householdbookapp/bnpl_delete.html'
    context_object_name = 'target_bnpl'

    def get_success_url(self):
        return reverse('householdbookapp:debt_list')


class BuyNowPayLaterUpdateView(UpdateView):
    model = BuyNowPayLater
    form_class = BuyNowPayLaterCreationForm
    template_name = 'householdbookapp/bnpl_update.html'
    context_object_name = 'target_bnpl'

    def get_success_url(self):
        return reverse('householdbookapp:debt_list')


class BuyNowPayLaterTotalListView(ListView):
    model = BuyNowPayLater
    template_name = 'householdbookapp/bnpl_list_total.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(BuyNowPayLaterTotalListView, self).get_context_data(**kwargs)

        queryset_my_dashboard = Dashboard.objects.get(owner=self.request.user)

        queryset_bnpls = BuyNowPayLater.objects.filter(owner=self.request.user,
                                                       dashboard=queryset_my_dashboard.pk).order_by('end_flag',
                                                                                                    'purchase_period')

        bnpl_summary = {
            'total_amount': 0,
            'discount_amount': 0,
            'real_remaining_amount': 0,
            'real_monthly_payment_amount': 0,
        }
        for bnpl in queryset_bnpls:
            bnpl.update_statistics()

            bnpl_summary['total_amount'] += bnpl.total_amount
            bnpl_summary['discount_amount'] += bnpl.discount_amount
            bnpl_summary['real_remaining_amount'] += bnpl.real_remaining_amount
            bnpl_summary['real_monthly_payment_amount'] += bnpl.real_monthly_payment_amount

            bnpl.payment_count = str(round(bnpl.current_payment_count))+' / '+str(round(bnpl.paying_months))
            bnpl.total_amount = format_mask_currency(bnpl.total_amount, bnpl.currency)
            bnpl.discount_amount = format_mask_currency(bnpl.discount_amount, bnpl.currency)
            bnpl.real_remaining_amount = format_mask_currency(bnpl.real_remaining_amount, bnpl.currency)
            bnpl.real_monthly_payment_amount = format_mask_currency(bnpl.real_monthly_payment_amount, bnpl.currency)
            bnpl.end_flag = 'Y' if bnpl.end_flag else 'N'

        context.update({'queryset_bnpls': queryset_bnpls})

        bnpl_summary['total_amount'] = format_mask_currency(bnpl_summary['total_amount'], queryset_my_dashboard.main_currency)
        bnpl_summary['discount_amount'] = format_mask_currency(bnpl_summary['discount_amount'], queryset_my_dashboard.main_currency)
        bnpl_summary['real_remaining_amount'] = format_mask_currency(bnpl_summary['real_remaining_amount'], queryset_my_dashboard.main_currency)
        bnpl_summary['real_monthly_payment_amount'] = format_mask_currency(bnpl_summary['real_monthly_payment_amount'], queryset_my_dashboard.main_currency)
        context.update({'bnpl_summary': bnpl_summary})

        return context