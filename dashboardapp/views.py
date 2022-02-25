import json
from datetime import datetime, timedelta

import pandas as pd
import pytz
from django import utils
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

# Create your views here.
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views.generic import CreateView, DetailView, ListView
from django.views.generic.edit import FormMixin, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from pyecharts.charts import Pie, Grid, Line, Page
from pyecharts import options as opts
from rest_framework.views import APIView

from assetapp.models import Asset, PensionAsset
from dashboardapp.decorators import dashboard_ownership_required
from dashboardapp.forms import DashboardCreationForm, AssetHistoryCreationForm, AssetHistoryCaptureForm
from dashboardapp.models import Dashboard, AssetHistory
from diamond_goose.pyecharts import json_response
from householdbookapp.models import Liquidity, Debt
from pensionapp.models import Pension
from portfolioapp.forms import PortfolioCreationForm
from portfolioapp.models import Portfolio

has_ownership = [login_required, dashboard_ownership_required]


class DashboardCreateView(CreateView):
    model = Dashboard
    form_class = DashboardCreationForm
    context_object_name = 'target_dashboard'
    template_name = 'dashboardapp/create.html'

    def form_valid(self, form):
        target_dashboard = form.save(commit=False)
        target_dashboard.owner = self.request.user
        target_dashboard.save()

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('dashboardapp:detail', kwargs={'pk':self.object.pk})


@method_decorator(has_ownership, 'get')
class DashboardDetailView(DetailView, FormMixin):
    model = Dashboard
    context_object_name = 'target_dashboard'
    form_class = PortfolioCreationForm
    template_name = 'dashboardapp/detail.html'

    def get_context_data(self, **kwargs):
        context = super(DashboardDetailView, self).get_context_data(**kwargs)

        initial_date = self.object.initial_date.replace(tzinfo=None)
        today = datetime.today()
        time_diff = today-initial_date
        d_day_count = time_diff.days
        context.update({'d_day_count': d_day_count})

        total_liquidity_amount = 0
        total_asset_amount = 0
        try:
            queryset_my_portfolio = Portfolio.objects.get(owner=self.request.user,
                                                          dashboard=self.object.pk)
            context.update({'target_portfolio_pk': queryset_my_portfolio.pk})

            queryset_my_liquidities = Liquidity.objects.filter(owner=self.request.user,
                                                               dashboard=self.object.pk)
            for liquidity in queryset_my_liquidities:
                total_liquidity_amount += liquidity.amount_exchanged
            total_asset_amount += queryset_my_portfolio.current_value + total_liquidity_amount

        except Exception as identifier:
            print('Dashboard Detail my_portfolio query :', identifier)

        queryset_my_debts = Debt.objects.filter(owner=self.request.user,
                                                dashboard=self.object.pk)
        long_term_debt_amount = 0
        short_term_debt_amount = 0
        for debt in queryset_my_debts:
            if debt.long_term_debt_flag:
                long_term_debt_amount += debt.amount_exchanged
            else:
                short_term_debt_amount += debt.amount_exchanged
        total_debt_amount = long_term_debt_amount + short_term_debt_amount
        net_capital_amount = total_asset_amount - total_debt_amount

        context.update({
            'date_today': datetime.today().strftime('%Y.%m.%d'),
            'total_asset_amount': self.set_format_mask(total_asset_amount),
            'net_capital_amount': self.set_format_mask(net_capital_amount),
            'total_debt_amount': self.set_format_mask(total_debt_amount),
        })


        asset_summary_pie_chart_url_list = ['http://']
        ip_address = None
        try:
            from diamond_goose.settings.local import LOCAL_IP_ADDRESS
            ip_address = LOCAL_IP_ADDRESS
        except Exception as deploy_environment:
            print('Liquidity Pie Chart - deploy_environment : {}'.format(deploy_environment))
            from diamond_goose.settings.deploy import DEPLOY_IP_ADDRESS
            ip_address = DEPLOY_IP_ADDRESS

        if ip_address:
            asset_summary_pie_chart_url_list.append(ip_address)
            asset_summary_pie_chart_url_list.append('/dashboard/detail_asset_summary_pie_chart/')
            context.update({'asset_summary_pie_chart_url_list': ''.join(asset_summary_pie_chart_url_list)})

        return context

    def set_format_mask(self, amount):
        main_currency = self.object.main_currency
        result_text_list = [main_currency.currency_sign, ' ']
        below_period = None
        if main_currency.currency_code != 'KRW':
            below_period = str(round(amount, 2)).split('.')[-1]

        integer_list = []
        while amount >= 1000:
            temp_num_str = str(round(amount%1000))
            while len(temp_num_str) < 3:
                temp_num_str = '0' + temp_num_str
            integer_list.append(temp_num_str)
            amount /= 1000
        integer_list.append(str(round(amount)))

        for i in range(len(integer_list)):
            result_text_list.append(integer_list[(i+1)*(-1)])
            result_text_list.append(',')
        result_text_list.pop(-1)
        if below_period:
            result_text_list.append('.')
            result_text_list.append(below_period)
        return ''.join(result_text_list)


def asset_summary_pie_chart_data_generator(request, dashboard_pk):
    large_x_data = []
    large_y_data = []
    large_color_list = []
    total_asset_amount = 0

    queryset_liquidity = Liquidity.objects.filter(dashboard=dashboard_pk)
    for liquidity in queryset_liquidity:
        large_x_data.append(liquidity.liquidity_name)
        large_y_data.append(liquidity.amount_exchanged)
        large_color_list.append("#068CD6")
        total_asset_amount += liquidity.amount_exchanged

    queryset_portfolio = Portfolio.objects.get(dashboard=dashboard_pk)
    queryset_my_assets = Asset.objects.filter(portfolio=queryset_portfolio.pk,
                                              position_opened_flag=True,
                                              quantity__gt=0).exclude(asset_master__asset_type_master__asset_type_code='PENSION_ASSET').order_by('asset_master__asset_type')
    for asset in queryset_my_assets:
        large_x_data.append(asset.asset_master.name)
        large_y_data.append(asset.total_amount_exchanged)
        large_color_list.append(asset.asset_master.asset_type_master.color_hex)
        total_asset_amount += asset.total_amount_exchanged

    queryset_my_pensions = Pension.objects.filter(portfolio=queryset_portfolio.pk)
    for pension in queryset_my_pensions:
        queryset_my_pension_assets = PensionAsset.objects.filter(pension=pension.pk)
        for pension_asset in queryset_my_pension_assets:
            if pension_asset.position_opened_flag:
                large_x_data.append(pension_asset.asset_master.name)
                large_y_data.append(pension_asset.total_amount_exchanged)
                pension_color = pension_asset.asset_master.asset_type_master.color_hex
                large_color_list.append(pension_color)
        large_x_data.append(pension.pension_master.pension_name+' Cash')
        large_y_data.append(pension.total_cash_amount)
        large_color_list.append(pension_color)
    large_data_pair = [list(z) for z in zip(large_x_data, large_y_data)]

    small_x_data = []
    small_y_data = []
    small_color_list = []
    total_debt_amount = 0
    queryset_my_debts = Debt.objects.filter(dashboard=dashboard_pk)
    for debt in queryset_my_debts:
        small_x_data.append(debt.debt_name)
        small_y_data.append(debt.amount_exchanged)
        if debt.long_term_debt_flag:
            small_color_list.append("#FA0067")
        else:
            small_color_list.append("#FA0067")
        total_debt_amount += debt.amount_exchanged
    small_x_data.append('Net Capital')
    small_y_data.append(total_asset_amount-total_debt_amount)
    small_color_list = ["#17344A"]
    small_data_pair = [list(z) for z in zip(small_x_data, small_y_data)]

    leverage_rate = 0
    if total_asset_amount > 0:
        leverage_rate = total_debt_amount/total_asset_amount

    return {
        'large_data_pair': large_data_pair,
        'large_color_list': large_color_list,
        'small_data_pair': small_data_pair,
        'small_color_list': small_color_list,
        'leverage_rate': str(round(leverage_rate*100, 2))+'%',
    }


def asset_summary_pie_chart(request, dump_option=False) -> Pie:
    queryset_my_dashboard = Dashboard.objects.get(owner=request.user)

    # Pie Graph Data
    chart_base_data = asset_summary_pie_chart_data_generator(request, queryset_my_dashboard.pk)
    large_data_pair = chart_base_data['large_data_pair']
    large_color_list = chart_base_data['large_color_list']
    small_data_pair = chart_base_data['small_data_pair']
    small_color_list = chart_base_data['small_color_list']
    leverage_rate = chart_base_data['leverage_rate']

    # AssetHistory Line Graph
    line_base_data = asset_history_line_graph_data_generator(request, queryset_my_dashboard.pk)
    line_x_data = line_base_data['x_data']
    line_y_data = line_base_data['line_y_data']
    max_y_value = line_base_data['max_y_value']

    line_graph = Line()
    line_graph.add_xaxis(xaxis_data=line_x_data)
    color_list = [
        '#068CD6',
        '#00C484',
        '#007D8A',
        '#FF9C00',
        '#FA0067',
    ]
    color_idx = 0
    for statistic in line_y_data:
        line_graph.add_yaxis(
            series_name=statistic['name'],
            y_axis=statistic['data_set'],
            is_symbol_show=False,
            linestyle_opts=opts.LineStyleOpts(type_='solid', width=2, color=color_list[color_idx]),
            label_opts=opts.LabelOpts(color=color_list[color_idx], font_weight='bold')
        )
        color_idx += 1
    line_graph.set_colors(
        large_color_list
    )
    line_graph.set_global_opts(
        title_opts=opts.TitleOpts(title="Asset History",
                                  title_textstyle_opts=opts.TextStyleOpts(color='#FFFFFF', font_size=25),
                                  pos_left="38%"),
        xaxis_opts=opts.AxisOpts(type_='category',
                                 boundary_gap=False,
                                 axislabel_opts=opts.LabelOpts(margin=20, color="#FFFFFF", rotate=15),
                                 axisline_opts=opts.AxisLineOpts(linestyle_opts=opts.LineStyleOpts(color='#FFFFFF')),
                                 axistick_opts=opts.AxisTickOpts(
                                     is_show=True,
                                     length=6,
                                     linestyle_opts=opts.LineStyleOpts(color="#FFFFFF")),
                                 ),
        yaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(color='#FFFFFF'),
                                 axisline_opts=opts.AxisLineOpts(linestyle_opts=opts.LineStyleOpts(color='#FFFFFF')),
                                 splitline_opts=opts.SplitLineOpts(
                                     is_show=True,
                                     linestyle_opts=opts.LineStyleOpts(color="#FFFFFF1F")),
                                 axispointer_opts=opts.AxisPointerOpts(is_show=True,
                                                                       label=opts.LabelOpts(color="#081321",
                                                                                            font_weight='bold')),
                                 position="right",
                                 ),
        tooltip_opts=opts.TooltipOpts(background_color='#FFFFFF', textstyle_opts=opts.TextStyleOpts(color="#081321")),
        legend_opts=opts.LegendOpts(is_show=True,
                                    pos_left='55%',
                                    pos_top='2%',
                                    textstyle_opts=opts.TextStyleOpts(color='#FFFFFF'),
                                    legend_icon=False,),
    )
    dump_line_graph = line_graph.dump_options_with_quotes()

    # Pie Graphs
    large_pie_chart = Pie()
    large_pie_chart.add(
            series_name="Asset Composition",
            data_pair=large_data_pair,
            radius=["40%", "70%"],
            label_opts=opts.LabelOpts(is_show=True, position="center"),
            itemstyle_opts=opts.ItemStyleOpts(border_color="#081321", border_width=1),
            center=["21%", "50%"],
        )
    large_pie_chart.set_colors(
            large_color_list
        )
    large_pie_chart.set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(color="#FFFFFF"),
        )
    large_pie_chart.set_global_opts(
        title_opts=opts.TitleOpts(title='Leverage : '+leverage_rate,
                                  pos_top="40%",
                                  pos_left="17%",
                                  title_textstyle_opts=opts.TextStyleOpts(color="#FA0067",
                                                                          font_size=15)),
        legend_opts=opts.LegendOpts(is_show=False),
    )

    for i in range(2):
        large_color_list.pop(-1)
    large_color_list.extend(small_color_list)
    small_pie_chart = Pie()
    small_pie_chart.add(
            series_name="Leverage Status",
            data_pair=small_data_pair,
            radius=["30%", "40%"],
            label_opts=opts.LabelOpts(is_show=False,),
            itemstyle_opts=opts.ItemStyleOpts(border_color="#081321", border_width=1),
            center=["21%", "50%"],
        )
    small_pie_chart.set_colors(
        large_color_list
    )
    small_pie_chart.set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(is_show=False),
        )
    small_pie_chart.set_global_opts(legend_opts=opts.LegendOpts(is_show=False))

    grid = Grid()
    grid.add(
        chart=line_graph,
        grid_opts=opts.GridOpts(pos_top="10%",
                                pos_right="5%",
                                width="55%")
    )
    grid.add(
        chart=large_pie_chart,
        grid_opts=opts.GridOpts(pos_top="60%",
                                pos_left="15%")
    )
    grid.add(
        chart=small_pie_chart,
        grid_opts=opts.GridOpts(pos_top="60%",
                                pos_left="15%")
    )

    dump_grid = grid.dump_options_with_quotes()

    if dump_option:
        return dump_grid
    else:
        return grid


class AssetSummaryPieChartView(APIView):
    def get(self, request, *args, **kwargs):
        return json_response(json.loads(asset_summary_pie_chart(self.request, dump_option=True)))

class AssetHistoryListView(ListView):
    model = AssetHistory
    template_name = 'dashboardapp/asset_history_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssetHistoryListView, self).get_context_data(**kwargs)

        amount_column_width_px = 110
        table_width_px = 70*2 + 180 + amount_column_width_px*5
        queryset_my_dashboard = Dashboard.objects.get(owner=self.request.user)
        queryset_my_asset_history = AssetHistory.objects.filter(owner=self.request.user,
                                                                dashboard=queryset_my_dashboard.pk).order_by('-capture_date')
        for asset_history in queryset_my_asset_history:
            asset_history.capture_date = asset_history.capture_date.strftime("%Y-%m-%d %H:%M:%S")
        context.update({
            'table_width_px': str(table_width_px)+'px',
            'amount_column_width_px': str(amount_column_width_px)+'px',
            'dashboard_pk': queryset_my_dashboard.pk,
            'queryset_my_asset_history': queryset_my_asset_history,
        })

        return context


def asset_history_excel_download(request):
    if request.method == 'POST':
        dashboard_pk = request.POST['dashboard_pk']
    queryset_asset_history = AssetHistory.objects.filter(dashboard=dashboard_pk).order_by('-capture_date')

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "transactions"
        output_row_count = 1

        # Header Insert
        worksheet.cell(row=output_row_count, column=1, value="capture_date")
        worksheet.cell(row=output_row_count, column=2, value="total_asset_amount")
        worksheet.cell(row=output_row_count, column=3, value="total_liquidity_amount")
        worksheet.cell(row=output_row_count, column=4, value="total_investment_amount")
        worksheet.cell(row=output_row_count, column=5, value="total_debt_amount")
        worksheet.cell(row=output_row_count, column=6, value="net_capital_amount")

        # Line Insert
        for asset_history in queryset_asset_history:
            output_row_count += 1
            worksheet.cell(row=output_row_count, column=1, value=asset_history.capture_date.strftime("%Y-%m-%d"))
            worksheet.cell(row=output_row_count, column=2, value=asset_history.total_asset_amount)
            worksheet.cell(row=output_row_count, column=3, value=asset_history.total_liquidity_amount)
            worksheet.cell(row=output_row_count, column=4, value=asset_history.total_investment_amount)
            worksheet.cell(row=output_row_count, column=5, value=asset_history.total_debt_amount)
            worksheet.cell(row=output_row_count, column=6, value=asset_history.net_capital_amount)

        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=asset_history.xlsx'
        return response

    except Exception as identifier:
        print('asset_history_excel_download: excel_export', identifier)


class AssetHistoryExcelUploadButton(ListView):
    model = AssetHistory
    template_name = 'dashboardapp/asset_history_excel_upload.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssetHistoryExcelUploadButton, self).get_context_data(**kwargs)
        context.update({
            'user_id': self.request.user.id,
            'dashboard_pk': Dashboard.objects.get(owner=self.request.user).pk,
        })

        return context


def asset_history_excel_upload(request):
    try:
        if request.method == 'POST' and request.FILES['transaction_file']:
            db_frame = pd.read_excel(request.FILES['transaction_file'], sheet_name=0).fillna('')
            target_dashboard = Dashboard.objects.get(pk=request.POST['dashboard_pk'])
    except Exception as identifier:
        print('income_expense_excel_upload: excel_import', identifier)

    try:
        transaction_object_list = []
        for row in db_frame.itertuples():
            upload_format = {}
            upload_format['capture_date'] = row[1]
            upload_format['total_asset_amount'] = row[2]
            upload_format['total_liquidity_amount'] = row[3]
            upload_format['total_investment_amount'] = row[4]
            upload_format['total_debt_amount'] = row[5]
            upload_format['net_capital_amount'] = row[6]

            tz = pytz.timezone('Asia/Seoul')
            capture_date_raw = upload_format['capture_date'].to_pydatetime()
            capture_date_timezone = tz.localize(capture_date_raw, is_dst=None).astimezone(pytz.utc)

            print(upload_format)

            obj = AssetHistory.objects.create(
                owner=request.user,
                dashboard=target_dashboard,
                capture_date=capture_date_timezone,
                total_asset_amount=upload_format['total_asset_amount'],
                total_liquidity_amount=upload_format['total_liquidity_amount'],
                total_investment_amount=upload_format['total_investment_amount'],
                total_debt_amount=upload_format['total_debt_amount'],
                net_capital_amount=upload_format['net_capital_amount'],
            )

            transaction_object_list.append(obj)

        # Transaction COMMIT
        for transaction in transaction_object_list:
            try:
                transaction.save()
                print('Transaction Added :', transaction)
            except Exception as identifier:
                print('Exception in saving transaction.', identifier, transaction)
                continue

    except Exception as identifier:
        print('asset_history_excel_upload: Transaction INSERT :', identifier)

    return HttpResponseRedirect(reverse('dashboardapp:asset_history_list'))


def asset_history_line_graph_data_generator(request, dashboard_pk):

    x_data = []
    line_y_data = [
        {'name': 'Asset', 'data_set': []},
        {'name': 'Liquidity', 'data_set': []},
        {'name': 'Investment', 'data_set': []},
        {'name': 'Net Capital', 'data_set': []},
        {'name': 'Dept', 'data_set': []},
    ]

    max_y_value = 0

    queryset_asset_history = AssetHistory.objects.filter(dashboard=dashboard_pk).order_by('capture_date')
    for asset_history in queryset_asset_history:
        x_data.append(asset_history.capture_date.strftime("%Y-%m-%d"))

        max_y_value = max(
            max_y_value,
            max([
                asset_history.total_asset_amount,
                asset_history.total_liquidity_amount,
                asset_history.total_investment_amount,
                asset_history.total_debt_amount,
                asset_history.net_capital_amount,
            ]))

        for statistic in line_y_data:
            if statistic['name'] =='Asset'           :     statistic['data_set'].append(asset_history.total_asset_amount)
            if statistic['name'] =='Liquidity'       :     statistic['data_set'].append(asset_history.total_liquidity_amount)
            if statistic['name'] =='Investment'      :     statistic['data_set'].append(asset_history.total_investment_amount)
            if statistic['name'] =='Net Capital'     :     statistic['data_set'].append(asset_history.net_capital_amount)
            if statistic['name'] =='Dept'            :     statistic['data_set'].append(asset_history.total_debt_amount)


    # Current Data Set
    today = datetime.today().strftime("%Y-%m-%d")
    queryset_my_liquidities = Liquidity.objects.filter(dashboard=dashboard_pk)
    queryset_my_debts = Debt.objects.filter(dashboard=dashboard_pk)
    queryset_my_portfolio = Portfolio.objects.get(dashboard=dashboard_pk)

    current_liquidity_amount = 0
    for liquidity in queryset_my_liquidities:
        current_liquidity_amount += liquidity.amount_exchanged

    current_debt_amount = 0
    for debt in queryset_my_debts:
        current_debt_amount += debt.amount_exchanged

    current_investment_amount = queryset_my_portfolio.current_value
    current_total_asset_amount = current_liquidity_amount + current_investment_amount
    current_net_capital = current_total_asset_amount - current_debt_amount

    x_data.append(today)
    for statistic in line_y_data:
        if statistic['name'] == 'Asset'      :     statistic['data_set'].append(current_total_asset_amount)
        if statistic['name'] == 'Liquidity'  :     statistic['data_set'].append(current_liquidity_amount)
        if statistic['name'] == 'Investment' :     statistic['data_set'].append(current_investment_amount)
        if statistic['name'] == 'Net Capital':     statistic['data_set'].append(current_net_capital)
        if statistic['name'] == 'Dept'       :     statistic['data_set'].append(current_debt_amount)

    return {
        'x_data': x_data,
        'line_y_data': line_y_data,
        'max_y_value': round(max_y_value * 1.1),
    }


class AssetHistoryUpdateView(UpdateView):
    model = AssetHistory
    form_class = AssetHistoryCreationForm
    template_name = 'dashboardapp/asset_history_update.html'
    context_object_name = 'target_asset_history'

    def get_success_url(self):
        return reverse('dashboardapp:asset_history_list')


class AssetHistoryDeleteView(DeleteView):
    model = AssetHistory
    template_name = 'dashboardapp/asset_history_delete.html'
    context_object_name = 'target_asset_history'

    def get_success_url(self):
        return reverse('dashboardapp:asset_history_list')


def asset_history_delete_all(request):
    dashboard_pk = request.GET['dashboard_pk']
    queryset_asset_history = AssetHistory.objects.filter(owner=request.user,
                                                         dashboard=dashboard_pk)
    delete_count = 0
    deleted_periods_list = []
    for asset_history in queryset_asset_history:
        delete_count += 1
        deleted_periods_list.append(asset_history.capture_date.strftime(" %Y-%m-%d"))
        asset_history.delete()
    print('Delete Income/Expense : {}'.format(','.join(deleted_periods_list)))

    return HttpResponseRedirect(reverse('dashboardapp:asset_history_list'))


def asset_history_capture(request):

    my_dashboard = Dashboard.objects.get(owner=request.user)

    tz = pytz.timezone('Asia/Seoul')
    yesterday_raw = datetime.today()-timedelta(2)
    yesterday = tz.localize(yesterday_raw, is_dst=None).astimezone(pytz.utc)
    today = tz.localize(datetime.today(), is_dst=None).astimezone(pytz.utc)

    try:
        existing_asset_history = AssetHistory.objects.filter(capture_date__gt=(yesterday))
    except Exception as existing_asset_history_search:
        print('asset_history_capture - existing_asset_history : {}'.format(existing_asset_history_search))

    if existing_asset_history:
        for asset_history in existing_asset_history:
            if asset_history.capture_date.strftime("%Y%m%d") == today.strftime("%Y%m%d"):
                asset_history.delete()
                print('Asset History Delete : {} / {}'.format(asset_history.capture_date,
                                                              asset_history.total_asset_amount))

    queryset_my_liquidities = Liquidity.objects.filter(dashboard=my_dashboard.pk)
    total_liquidity_amount = 0
    for liquidity in queryset_my_liquidities:
        total_liquidity_amount += liquidity.amount_exchanged

    total_investment_amount = 0
    try:
        my_portfolio = Portfolio.objects.get(dashboard=my_dashboard.pk)
        my_portfolio.update_statistics(price_update=True)
        total_investment_amount += my_portfolio.current_value
    except Exception as my_portfolio_search:
        print('asset_history_capture - my_portfolio : {}'.format(my_portfolio_search))

    queryset_my_debts = Debt.objects.filter(dashboard=my_dashboard.pk)
    total_debt_amount = 0
    for debt in queryset_my_debts:
        total_debt_amount += debt.amount_exchanged

    temp_asset_history = AssetHistory.objects.create(
        owner=request.user,
        dashboard=my_dashboard,
        capture_date=today,
        total_asset_amount=(total_liquidity_amount+total_investment_amount),
        total_liquidity_amount=total_liquidity_amount,
        total_investment_amount=total_investment_amount,
        total_debt_amount=total_debt_amount,
        net_capital_amount=(total_liquidity_amount+total_investment_amount-total_debt_amount),
    )
    temp_asset_history.save()

    return HttpResponseRedirect(reverse('dashboardapp:detail', kwargs={'pk': my_dashboard.pk}))


